from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "mastersheet.xlsx"
OUTPUT = ROOT / "data.js"
FISH_IMAGE_DIR = ROOT / "assets" / "fish"

WATER_RU = {
    "Freshwater": "Пресная вода",
    "Saltwater": "Соленая вода",
}

FISH_RU = {
    "Belphoret Bass": "Бельфоретский окунь",
    "Silver Anchovy": "Серебристый анчоус",
    "Largemouth Bass": "Большеротый окунь",
    "Pink Cod": "Жёлтая треска",
    "Banded Butterflyfish": "Полосатая рыба-бабочка",
    "Arowana": "Аравана",
    "Taion Moray Eel": "Тайонская мурена",
    "Fonsine Barracuda": "Фонсинелова щука",
    "Ballan Wrasse": "Балланский губан",
    "Dolly Varden Trout": "Долливарнская форель",
    "Azure fishes (10 types)": "Лазурные рыбы (10 видов)",
    "Lava Starfish": "Лавовая звезда",
    "Leaf Sea Horse": "Листовой морской конёк",
    "Amberjack": "Коронада",
    "Pajama Cardinalfish": "Чёрнополосный апогон",
    "Imp Catalufa": "Карликовая каталуфа",
    "Flame Fish": "Огненная рыба",
    "Barracuda": "Барракуда",
    "Ruby Phantasmal Fish": "Рубиновая призрачная рыба",
    "Amber Phantasmal Fish": "Янтарная призрачная рыба",
    "Topaz Phantasmal Fish": "Топазовая призрачная рыба",
    "Emerald Phantasmal Fish": "Изумрудная призрачная рыба",
    "Crystal Phantasmal Fish": "Кристальная призрачная рыба",
    "Sapphire Phantasmal Fish": "Сапфировая призрачная рыба",
    "Lavender Phantasmal Fish": "Лавандовая призрачная рыба",
    "Rose Phantasmal Fish": "Розовая призрачная рыба",
    "Lanquis Barracuda": "Лангисова щука",
    "Ghost Thornfish": "Призрачный шипорыб",
    "Lava Catfish": "Лавовый сом",
    "Aurora Starfish": "Золотая морская звезда",
    "Solisium Carp": "Солизиумский карп",
    "Striped Mackerel": "Полосатая скумбрия",
    "Black Bream": "Черный лещ",
    "Small-Mouthed Rockfish": "Малоротый морской окунь",
    "Squid": "Кальмар",
    "Rainbow Ray-Finned Fish": "Радужная лучеперая рыба",
    "Obsidian Starfish": "Обсидиановая морская звезда",
    "Luderick": "Людерик",
    "Bluegill": "Синежаберник",
    "Lava Eel": "Лавовый угорь",
    "Blacktail Snook": "Чернохвостый снук",
    "Blue filefish": "Синий файлефиш",
    "Mackerel": "Скумбрия",
    "Lanquis Carp": "Карп Ланкис",
    "Mullet": "Кефаль",
    "Swamp Eel": "Болотный угорь",
    "Swamp Mudfish": "Болотная илистая рыба",
    "Mahi Mahi": "Махи-махи",
    "Horse Mackerel": "Ставрида",
    "Ice Smelt": "Ледяная корюшка",
    "Skipjack Tuna": "Тунец-скипджек",
    "Mantis Shrimp": "Креветка-богомол",
    "Discus": "Дискус",
    "Sunspot Rockfish": "Солнечно-пятнистый морской окунь",
    "Mushroom Jellyfish": "Грибная медуза",
    "Striped Crimson Seabream": "Полосатый багровый морской карась",
    "Pennant Coralfish": "Вымпельная коралловая рыба",
    "Heliber Goldfish": "Золотая рыбка Хелибер",
    "Red Salmon": "Красный лосось",
    "Peacock Bass": "Павлиний окунь",
    "Blacktip Shark": "Черноперая акула",
    "Kahawai": "Кахавай",
    "Sea Cockatrice": "Морской кокатрикс",
    "Lanquis Bass": "Бас Ланкис",
}

LOCATION_RU = {
    "Freshwater": "Пресная вода",
    "Saltwater": "Соленая вода",
    "Kastleton": "Каслтон",
    "Golden Rye Pastures": "Золотые нивы",
    "Blackhowl Plains": "Воющие равнины",
    "Urstella Fields": "Урстеллийские поля",
    "Carmine Forest": "Карминовый лес",
    "Nesting Grounds": "Дикие гнездовья",
    "Windhill Shores": "Ветреная отмель",
    "Daybreak Shore": "Рассветный берег",
    "Laslan Coast": "Побережье Ласлана",
    "Vienta Village": "Виента",
    "Veder Mountains": "Горы Ведер",
    "Sanctuary Oasis": "Священный оазис",
    "Moonlight Desert": "Лунная пустыня",
    "Syleus's Abyss Entrance": "Вход в Бездну Силеуса",
    "Syleus's Abyss 4F": "Бездна Силеуса, 4 этаж",
    "Syleus's Abyss 5F": "Бездна Силеуса, 5 этаж",
    "Syleus's Abyss 6F": "Бездна Силеуса, 6 этаж",
    "Watcher's Post": "Пост Смотрителя",
    "Akidu Valley": "Долина Акиду",
    "Stoneguard Coast": "Побережье Стоунгарда",
    "Stonegard Coast": "Побережье Стоунгарда",
    "Purelight Tower": "Светозарная башня",
    "Purelight Hill": "Светозарный холм",
    "Ruins of Turayne": "Развалины Турайна",
    "Tevent Temple (raining)": "Храм Тевента (дождь)",
    "Fonos Basin": "Фоносская котловина",
    "Stoneguard Castle (raining)": "Замок Стоунгард (дождь)",
    "Hall of Illusion": "Зал иллюзий",
    "Sanctum of Desire B1": "Святилище желания B1",
    "Sanctum of Desire Depths": "Глубины Святилища желания",
    "Quietis's Demesne": "Владения Квиетиса",
    "Forest of The Great Tree": "Лес Великого дерева",
    "Deluzhnoa Temple": "Храм Делузноа",
    "Bercant Manor": "Поместье Беркант",
    "Akidu Lake (needs rain)": "Озеро Акиду (нужен дождь)",
    "Herba Village": "Деревня Херба",
    "Swamp of Silence": "Болото тишины",
    "Black Anvil Forge": "Кузница Черной наковальни",
    "Temple of Truth B1": "Храм истины B1",
    "Temple of Truth Entrance": "Вход в Храм истины",
    "Red Fog Island": "Остров красного тумана",
    "Crimson Manor": "Багровое поместье",
    "Crimson Mansion B2": "Багровый особняк B2",
    "Grayclaw Forest": "Лес Серого когтя",
    "Greyclaw Forest": "Лес Серого когтя",
    "Saurodoma Island": "Остров Сауродома",
}

NOTE_RU = {
    "They have been present in the Summer 2025 event.\nMay appear again in future events.": "Были доступны в летнем событии 2025 года. Могут появиться снова в будущих событиях.",
    "Famous Fishing Spot required": "Требуется знаменитое место рыбалки.",
    "Halls of illusion secret room, chance to appear after killing first boss": "Секретная комната Зала иллюзий, шанс появления после убийства первого босса.",
    "Famous Fishing Spot required\n(For Ruins location look at the picture on the right)": "Требуется знаменитое место рыбалки. Для локации в Развалинах смотри подсказку в файле.",
    "Night": "Ночь.",
    "Only via Famous Fishing Spot Event": "Только через событие знаменитого места рыбалки.",
    "Rain": "Дождь.",
}

TIP_RU = {
    'You can speed up your fishing by utilising an animation cancel, when you have caught your fish press "ctrl+f" twice, this will put your rod away and redraw it skipping the "hold fish animation"': 'Можно ускорить рыбалку отменой анимации: когда рыба поймана, нажми "Ctrl+F" дважды, чтобы убрать и снова достать удочку, пропустив анимацию удержания рыбы.',
    "Your Fishing Level is based off of level + rod boost.\ne.g. with level 8 fishing and a +4 fishing rod it will display as level 12.": "Уровень рыбалки считается как уровень персонажа плюс бонус удочки. Например: уровень рыбалки 8 и удочка +4 покажут уровень 12.",
    "Using Paste bait increases chances of small + big pouches when fishing (rewards natures jade)": "Paste Bait повышает шанс получить малые и большие мешочки при рыбалке, из которых можно получить Nature's Jade.",
    "Fishing hotspots spawn every 3 hours, the time varies per server/region": "Рыболовные hotspots появляются каждые 3 часа, время зависит от сервера и региона.",
    'Settings - Content - "Fix Camera while Fishing" (personal preference) - i prefer it off': 'Настройка Content -> "Fix Camera while Fishing" зависит от предпочтений; в источнике советуют выключать.',
    'Sort your fishing log by "Lv." and begin working down the list': 'Отсортируй журнал рыбалки по "Lv." и проходи список сверху вниз.',
    "Your chance to fail whilst fishing significantly decreses as you level up": "Шанс неудачи при рыбалке заметно снижается с ростом уровня.",
    "Your best bet when aiming to level up your fishing is to get a bamboo rod and start \ncompleting the fishing log and catching everything you can at your level whilst moving\nspots. Whilst doing this you should send amitoi expeditions in Talandre if 20+\n(if not do Daybreak Shore) and always use your mystic keys in Talandre.": "Для прокачки лучше взять Bamboo Rod, закрывать журнал рыбалки и ловить всё доступное на твоём уровне, переходя между местами. Параллельно отправляй Amitoi экспедиции в Talandre, если уровень 20+, иначе в Daybreak Shore, и используй Mystic Keys в Talandre.",
}

BAIT_RU = {
    "Paste Bait": "Пастообразная наживка",
    "Fishing Bait Chest - Contract Coin Merchant- 20 coins for 50 bait, daily": "Fishing Bait Chest у Contract Coin Merchant: 20 монет за 50 наживки, ежедневно.",
    "Worm Bait": "Черви",
    "Cutting Tree's": "Рубка деревьев.",
    "Shrimp Bait": "Креветочная наживка",
    "Chestacean, Hermit Crab, Hermit Lobster, Iron Chestacean, Large Stone Crab": "Chestacean, Hermit Crab, Hermit Lobster, Iron Chestacean, Large Stone Crab.",
    "Fish Fillet Bait": "Рыбное филе",
    "Dissolving Fish": "Разбор рыбы.",
}

ROD_RU = {
    "Bamboo": "Бамбуковая",
    "Steel": "Стальная",
    "Monster Bone": "Из кости монстра",
    "Platinum": "Платиновая",
    "Kraken": "Кракен",
    "Tevent": "Тевент",
    "Obsidian": "Обсидиановая",
    "Luminous Dreamcatcher": "Светящийся ловец снов",
    "Whisperion Fishing Rod": "Удочка Whisperion",
    "Quest - Adventures of the Expedition": "Квест Adventures of the Expedition.",
    "Crafting - 10 Nature's Jade": "Крафт: 10 Nature's Jade.",
    "Crafting - 30 Nature's Jade": "Крафт: 30 Nature's Jade.",
    "Mystic Portals - Laslan/Stoneguard": "Mystic Portals: Laslan / Stoneguard.",
    "Amitoi Expedition - Daybreak Shore": "Amitoi Expedition: Daybreak Shore.",
    "Tevent Reward Chest / Crafting: 1 Tevent Soul + 100 Nature's Jade": "Tevent Reward Chest или крафт: 1 Tevent Soul + 100 Nature's Jade.",
    "Crafting: 1 Monster Bone Rod + 60 Nature's Jade + 30 Obsidian Piece": "Крафт: 1 Monster Bone Rod + 60 Nature's Jade + 30 Obsidian Piece.",
    "Crafting: 15 Dream Jade": "Крафт: 15 Dream Jade.",
    "Amitoi Expedition - Talandre\nMystic Portal/Globes - Talandre": "Amitoi Expedition: Talandre. Mystic Portal / Globes: Talandre.",
}


def ru(mapping: dict[str, str], value: str) -> str:
    return mapping.get(value, value)


def slugify(value: str) -> str:
    value = value.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "fish"


def extract_fish_images(ws) -> dict[int, str]:
    FISH_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    image_by_row = {}

    for image in getattr(ws, "_images", []):
        marker = getattr(image.anchor, "_from", None)
        if marker is None:
            continue

        row = marker.row + 1
        col = marker.col + 1
        fish_name = ws.cell(row, 3).value or ws.cell(row, 2).value

        if col != 2 or not fish_name:
            continue

        fish_en = str(fish_name).replace("\n", " ").strip()
        extension = (getattr(image, "format", None) or "png").lower()
        if extension == "jpeg":
            extension = "jpg"

        filename = f"{slugify(fish_en)}.{extension}"
        output_path = FISH_IMAGE_DIR / filename
        output_path.write_bytes(image._data())
        image_by_row[row] = f"assets/fish/{filename}"

    return image_by_row


def text_pair(value: str, mapping: dict[str, str]) -> dict[str, str]:
    clean = str(value).strip()
    return {"en": clean, "ru": ru(mapping, clean)}


def extract_info(ws) -> dict[str, list[dict[str, object]]]:
    tips = []
    for row in range(16, 25):
        value = ws.cell(row, 9).value
        if value:
            tips.append(text_pair(value, TIP_RU))

    baits = []
    for row in range(27, 31):
        bait = ws.cell(row, 9).value
        source = ws.cell(row, 10).value
        if bait and source:
            baits.append({
                "name": text_pair(bait, BAIT_RU),
                "source": text_pair(source, BAIT_RU),
            })

    rods = []
    for row in range(33, 42):
        rod = ws.cell(row, 9).value
        source = ws.cell(row, 10).value
        boost = ws.cell(row, 11).value
        if rod and source:
            rods.append({
                "name": text_pair(rod, ROD_RU),
                "source": text_pair(source, ROD_RU),
                "boost": int(boost) if isinstance(boost, float) and boost.is_integer() else boost,
            })

    return {"tips": tips, "baits": baits, "rods": rods}


def main() -> None:
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb["Master Sheet"]
    image_by_row = extract_fish_images(ws)
    fish_rows = []
    current = None

    for row in range(2, ws.max_row + 1):
        fish_name = ws.cell(row, 3).value or ws.cell(row, 2).value
        level = ws.cell(row, 4).value
        water = ws.cell(row, 5).value
        habitat = ws.cell(row, 6).value

        if fish_name:
            fish_en = str(fish_name).replace("\n", " ").strip()
            water_en = str(water).strip() if water else ""
            note = ws.cell(row, 8).value
            current = {
                "id": f"fish-{len(fish_rows) + 1}",
                "name": {"en": fish_en, "ru": ru(FISH_RU, fish_en)},
                "level": int(level) if isinstance(level, float) and level.is_integer() else level,
                "waterType": {"en": water_en, "ru": ru(WATER_RU, water_en)},
                "image": image_by_row.get(row),
                "note": text_pair(note, NOTE_RU) if note else None,
                "locations": [],
            }
            fish_rows.append(current)

        if current and habitat:
            location_en = str(habitat).strip()
            location = {"en": location_en, "ru": ru(LOCATION_RU, location_en)}
            if location not in current["locations"]:
                current["locations"].append(location)

    payload = {
        "source": "T&L Fishing Mastersheet.xlsx",
        "generatedFromRows": ws.max_row,
        "info": extract_info(ws),
        "fish": fish_rows,
    }
    OUTPUT.write_text(
        "window.FISHING_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT} with {len(fish_rows)} fish")


if __name__ == "__main__":
    main()
